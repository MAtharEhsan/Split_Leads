"""
Usage:
    python split_leads.py <excel_file> <leads_per_sheet>

Reads all leads from the workbook, removes duplicates, splits the unique
leads into sheets of size <leads_per_sheet>, and prints duplicate rows at
the end.
"""

import os
import re
import sys

from openpyxl import Workbook, load_workbook


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def clean_phone(value):
    if value is None:
        return ""
    return re.sub(r"\D", "", str(value))


def duplicate_key(headers, row):
    header_lookup = [clean_text(header) for header in headers]
    phone_index = None

    for index, header in enumerate(header_lookup):
        if "phone" in header or "mobile" in header or "contact" in header or "whatsapp" in header:
            phone_index = index
            break

    if phone_index is not None and phone_index < len(row):
        return clean_phone(row[phone_index])

    return tuple(clean_text(value) for value in row)


def main():
    print("\n" + "="*60)
    print("LEAD SPLITTER TOOL")
    print("="*60)
    print("\nThis tool splits leads from an Excel file into multiple")
    print("sheets with a specified number of leads per sheet.\n")

    filepath = input("Enter the path to the Excel file: ").strip()
    if not os.path.isfile(filepath):
        print(f"Error: file not found: {filepath}")
        sys.exit(1)

    while True:
        try:
            leads_per_sheet_input = input("\nEnter the number of leads required per sheet (positive integer): ").strip()
            leads_per_sheet = int(leads_per_sheet_input)
            if leads_per_sheet <= 0:
                print("Error: number must be greater than 0. Please try again.")
                continue
            break
        except ValueError:
            print("Error: invalid input. Please enter a positive integer.")
            continue

    print(f"\nProcessing: {filepath}")
    print(f"Leads per sheet: {leads_per_sheet}\n")

    workbook = load_workbook(filepath)

    headers = None
    unique_rows = []
    duplicates = []
    seen_keys = set()

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        if headers is None:
            headers = list(rows[0])

        for row_number, row in enumerate(rows[1:], start=2):
            values = list(row)
            if not any(clean_text(value) for value in values):
                continue

            key = duplicate_key(headers, values)
            if key in seen_keys:
                duplicates.append((sheet.title, row_number, values))
                continue

            seen_keys.add(key)
            unique_rows.append(values)

    if headers is None:
        print("Error: no data found in workbook.")
        sys.exit(1)

    batches = [unique_rows[index : index + leads_per_sheet] for index in range(0, len(unique_rows), leads_per_sheet)]
    if not batches:
        batches = [[]]

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)

    for batch_index, batch in enumerate(batches, start=1):
        worksheet = output_workbook.create_sheet(title=f"List {batch_index}")

        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

        for row_index, row in enumerate(batch, start=2):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

    output_path = filepath
    try:
        output_workbook.save(output_path)
    except PermissionError:
        base_name, extension = os.path.splitext(filepath)
        output_path = f"{base_name}_split{extension}"
        output_workbook.save(output_path)

    print(f"Saved file: {output_path}")
    print(f"Total rows loaded: {len(unique_rows) + len(duplicates)}")
    print(f"Unique leads: {len(unique_rows)}")
    print(f"Duplicates removed: {len(duplicates)}")
    print(f"Sheets created: {len(batches)}")

    if duplicates:
        print("\nDuplicate entries:")
        for sheet_name, row_number, values in duplicates:
            print(f"{sheet_name} row {row_number}: {values}")
    else:
        print("\nNo duplicates found.")


if __name__ == "__main__":
    main()
